"""
Universal Large File to CSV Converter
Handles: JSON, JSONL, SQL, XML, CSV, TXT, LOG, XLSX
Supports files 1 GB - 10+ GB via streaming/chunked processing
"""

import os
import sys
import csv
import json
import re
import time

# ─── CONFIGURATION ────────────────────────────────────────────────────────────
INPUT_FILE  = r"C:\PUT\YOUR\FILE\PATH\HERE.json"   # <-- change this
OUTPUT_DIR  = r"C:\Users\NB\OneDrive\เดสก์ท็อป\Ai_For_csv"
CHUNK_SIZE  = 50_000        # rows per output CSV file
ENCODING    = "utf-8"
DELIMITER   = ","
# ──────────────────────────────────────────────────────────────────────────────

def detect_format(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".json",):     return "json"
    if ext in (".jsonl", ".ndjson"): return "jsonl"
    if ext in (".sql",):      return "sql"
    if ext in (".xml",):      return "xml"
    if ext in (".xlsx", ".xls"): return "xlsx"
    if ext in (".csv", ".txt", ".log", ".tsv"): return "flat"
    return "flat"

def flatten(d, parent="", sep="_"):
    items = {}
    for k, v in d.items():
        key = f"{parent}{sep}{k}" if parent else k
        if isinstance(v, dict):
            items.update(flatten(v, key, sep))
        elif isinstance(v, list):
            items[key] = json.dumps(v, ensure_ascii=False)
        else:
            items[key] = v
    return items

def make_output_path(base, part):
    name = os.path.splitext(os.path.basename(base))[0]
    return os.path.join(OUTPUT_DIR, f"{name}_part{part:03d}.csv")

def write_chunk(rows, headers, out_path):
    with open(out_path, "w", newline="", encoding=ENCODING) as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Wrote {len(rows):,} rows → {out_path}")

# ── JSON (full array) ──────────────────────────────────────────────────────────
def convert_json(path):
    print("[JSON] Loading...")
    with open(path, "r", encoding=ENCODING, errors="replace") as f:
        data = json.load(f)
    if isinstance(data, dict):
        data = [data]
    rows = [flatten(r) if isinstance(r, dict) else {"value": r} for r in data]
    headers = list(dict.fromkeys(k for r in rows for k in r))
    _save_chunks(rows, headers, path)

# ── JSONL / NDJSON ─────────────────────────────────────────────────────────────
def convert_jsonl(path):
    print("[JSONL] Streaming line by line...")
    headers, buf, part, total = [], [], 1, 0
    seen = set()
    with open(path, "r", encoding=ENCODING, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = flatten(json.loads(line))
            except json.JSONDecodeError:
                continue
            for k in obj:
                if k not in seen:
                    seen.add(k); headers.append(k)
            buf.append(obj)
            total += 1
            if len(buf) >= CHUNK_SIZE:
                write_chunk(buf, headers, make_output_path(path, part))
                part += 1; buf = []
    if buf:
        write_chunk(buf, headers, make_output_path(path, part))
    print(f"[JSONL] Done. {total:,} rows total.")

# ── SQL INSERT ─────────────────────────────────────────────────────────────────
def convert_sql(path):
    print("[SQL] Extracting INSERT statements...")
    pattern = re.compile(
        r"INSERT\s+INTO\s+`?(\w+)`?\s*\(([^)]+)\)\s*VALUES\s*(.+?);",
        re.IGNORECASE | re.DOTALL
    )
    val_pattern = re.compile(r"\(([^()]*(?:'[^']*'[^()]*)*)\)")
    tables = {}
    buf_map = {}
    part_map = {}
    total = 0

    with open(path, "r", encoding=ENCODING, errors="replace") as f:
        content = ""
        for line in f:
            content += line
            if ";" in content:
                for m in pattern.finditer(content):
                    tbl = m.group(1)
                    cols = [c.strip().strip("`'\"") for c in m.group(2).split(",")]
                    if tbl not in tables:
                        tables[tbl] = cols
                        buf_map[tbl] = []
                        part_map[tbl] = 1
                    for vm in val_pattern.finditer(m.group(3)):
                        raw = vm.group(1)
                        vals = [v.strip().strip("'") for v in re.split(r",(?=(?:[^']*'[^']*')*[^']*$)", raw)]
                        row = dict(zip(cols, vals))
                        buf_map[tbl].append(row)
                        total += 1
                        if len(buf_map[tbl]) >= CHUNK_SIZE:
                            out = os.path.join(OUTPUT_DIR, f"{tbl}_part{part_map[tbl]:03d}.csv")
                            write_chunk(buf_map[tbl], tables[tbl], out)
                            part_map[tbl] += 1
                            buf_map[tbl] = []
                content = ""

    for tbl in tables:
        if buf_map[tbl]:
            out = os.path.join(OUTPUT_DIR, f"{tbl}_part{part_map[tbl]:03d}.csv")
            write_chunk(buf_map[tbl], tables[tbl], out)
    print(f"[SQL] Done. {total:,} rows across {len(tables)} table(s): {list(tables.keys())}")

# ── XML ────────────────────────────────────────────────────────────────────────
def convert_xml(path):
    try:
        import xml.etree.ElementTree as ET
    except ImportError:
        print("xml.etree.ElementTree not available"); return
    print("[XML] Streaming records...")
    buf, headers, part, total = [], [], 1, 0
    seen = set()
    context = ET.iterparse(path, events=("end",))
    root_tag = None
    for event, elem in context:
        if root_tag is None:
            root_tag = elem.tag
            continue
        if len(list(elem)) > 0 or elem.tag == root_tag:
            row = {child.tag: (child.text or "").strip() for child in elem}
            if not row:
                continue
            for k in row:
                if k not in seen:
                    seen.add(k); headers.append(k)
            buf.append(row); total += 1
            elem.clear()
            if len(buf) >= CHUNK_SIZE:
                write_chunk(buf, headers, make_output_path(path, part))
                part += 1; buf = []
    if buf:
        write_chunk(buf, headers, make_output_path(path, part))
    print(f"[XML] Done. {total:,} rows.")

# ── XLSX ───────────────────────────────────────────────────────────────────────
def convert_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl..."); os.system("pip install openpyxl")
        import openpyxl
    print("[XLSX] Reading sheets...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows))]
        buf, part, total = [], 1, 0
        for row in rows:
            buf.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
            total += 1
            if len(buf) >= CHUNK_SIZE:
                out = os.path.join(OUTPUT_DIR, f"{os.path.splitext(os.path.basename(path))[0]}_{sheet}_part{part:03d}.csv")
                write_chunk(buf, headers, out); part += 1; buf = []
        if buf:
            out = os.path.join(OUTPUT_DIR, f"{os.path.splitext(os.path.basename(path))[0]}_{sheet}_part{part:03d}.csv")
            write_chunk(buf, headers, out)
        print(f"  Sheet '{sheet}': {total:,} rows.")

# ── CSV / TXT / LOG ────────────────────────────────────────────────────────────
def convert_flat(path):
    print("[FLAT] Auto-detecting delimiter...")
    with open(path, "r", encoding=ENCODING, errors="replace") as sample:
        sniff = sample.read(4096)
    try:
        dialect = csv.Sniffer().sniff(sniff)
        delim = dialect.delimiter
    except Exception:
        delim = "\t" if "\t" in sniff else ","
    print(f"  Detected delimiter: {repr(delim)}")
    with open(path, "r", encoding=ENCODING, errors="replace") as f:
        reader = csv.DictReader(f, delimiter=delim)
        headers = reader.fieldnames or []
        buf, part, total = [], 1, 0
        for row in reader:
            buf.append(row); total += 1
            if len(buf) >= CHUNK_SIZE:
                write_chunk(buf, headers, make_output_path(path, part))
                part += 1; buf = []
        if buf:
            write_chunk(buf, headers, make_output_path(path, part))
    print(f"[FLAT] Done. {total:,} rows.")

def _save_chunks(rows, headers, path):
    for i in range(0, len(rows), CHUNK_SIZE):
        chunk = rows[i:i+CHUNK_SIZE]
        part = i // CHUNK_SIZE + 1
        write_chunk(chunk, headers, make_output_path(path, part))
    print(f"Done. {len(rows):,} rows total.")

# ── MAIN ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else INPUT_FILE
    if not os.path.exists(path):
        print(f"ERROR: File not found: {path}"); sys.exit(1)

    size_gb = os.path.getsize(path) / (1024**3)
    fmt = detect_format(path)
    print(f"\nFile : {path}")
    print(f"Size : {size_gb:.2f} GB")
    print(f"Format: {fmt.upper()}")
    print(f"Output: {OUTPUT_DIR}\n")

    start = time.time()
    if   fmt == "json":  convert_json(path)
    elif fmt == "jsonl": convert_jsonl(path)
    elif fmt == "sql":   convert_sql(path)
    elif fmt == "xml":   convert_xml(path)
    elif fmt == "xlsx":  convert_xlsx(path)
    else:                convert_flat(path)

    elapsed = time.time() - start
    print(f"\nCompleted in {elapsed:.1f} seconds.")
    print(f"CSV files saved to: {OUTPUT_DIR}")
